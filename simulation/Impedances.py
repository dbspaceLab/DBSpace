#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 19 16:39:29 2018

@author: virati
"""

import sys
#sys.path.append('/home/virati/Dropbox/projects/Research/MDD-DBS/Ephys/IntegratedAnalysis/')
sys.path.append('/home/virati/Dropbox/projects/Research/MDD-DBS/Ephys/DBSpace/')
#import DBSpace as dbs
import DBSpace as dbo
from DBSpace import nestdict

import ipdb

import itertools as itt

import numpy as np
import scipy.signal as sig
import matplotlib.pyplot as plt
import pandas as pd

import seaborn as sns

sns.set_context('paper')
sns.set(font_scale=2)
sns.set_style('white')

from openpyxl import load_workbook

# Create a function to calculat the total brain volume
def total_brain_volume(gm, wm, wsf):
    # returns the sum of gm, wm, wsf
    return gm + wm + wsf

# Create a function to calculat the total brain volume
def gm_total_ratio(gm, total):
    # returns the ratio of gm to total brain volume
    return gm / total        
        
class Anatomy:
    def __init__(self):
        lead = nestdict()
        
        lead['Left'] = pd.read_csv('/home/virati/Dropbox/projects/Research/MDD-DBS/Data/Anatomy/CT/VTRVTA_6AMP_1.csv',  names=['Subj', 'Contact', 'GM','WM', 'CSF'])
        lead['Right'] = pd.read_csv('/home/virati/Dropbox/projects/Research/MDD-DBS/Data/Anatomy/CT/VTRVTA_6AMP_2.csv',  names=['Subj', 'Contact', 'GM', 'WM','CSF'])
        
        for side in ['Left','Right']:
            # Create a total_volume variable that is the output of the function
            lead[side]['Total'] = total_brain_volume(lead[side]['GM'], lead[side]['WM'], lead[side]['CSF'])
            lead[side]['Gm_Total_Ratio'] = gm_total_ratio(lead[side]['GM'], lead[side]['Total'])


class Z_class:
    def __init__(self):
        #fname = '/home/virati/Downloads/MDT_Zs.xlsx'
        fname = '/home/virati/Dropbox/projects/Research/MDD-DBS/Data/Anatomy/CT/Zs_GMs_Es.xlsx'
        wb = load_workbook(fname)
        ws = wb['Zs']
        self.xs = nestdict()
        self.pts = ['DBS901','DBS903','DBS905','DBS906','DBS907','DBS908']
        for pp,pt in enumerate(self.pts):
            first_col = (pp)*5
            print(first_col)
            self.xs[pt]['Date'] = np.array([r[first_col].value for r in ws.iter_rows()])[1:]
            self.xs[pt]['Left'] = np.array([r[first_col+2].value for r in ws.iter_rows()])[1:]
            self.xs[pt]['Right'] = np.array([r[first_col+4].value for r in ws.iter_rows()])[1:]
            
            self.xs[pt]['Date'] = self.xs[pt]['Date'][self.xs[pt]['Date'] != None]
            
            #F ORDER IS CORRECT, gives us 4xobs matrix for each electrode
            self.xs[pt]['Left'] = self.xs[pt]['Left'].reshape(4,-1,order='F')
            self.xs[pt]['Right'] = self.xs[pt]['Right'].reshape(4,-1,order='F')
            #self.xs[pt]['Left'] = self.xs[pt]['Left'][self.xs[pt]['Left'] != None].reshape(4,-1,order='F')
            #self.xs[pt]['Right'] = self.xs[pt]['Right'][self.xs[pt]['Right'] != None].reshape(4,-1,order='F')
            
        self.get_recZs()
        self.gen_Zdiff()
            
    def ret_Z(self,pt,etrode):
        if etrode > 4:
            side = 'Right'
            etrode = etrode - 8
        else:
            side = 'Left'
            
        return self.xs['DBS'+pt][side][etrode::4]
    
    def load_pt_OnT_stim(self):
        pt_OnT = [(2,1),(2,2),(2,1),(2,2),(1,1),(2,1)]
        #self.pt_OnT = {pt:pt_OnT[pp] for pp,pt in enumerate(self.do_pts)}
        return pt_OnT
    
    def get_recZs(self):
        on_t_es = self.load_pt_OnT_stim()
        
        rec_Zs = {'Left':np.zeros((6,2,28)),'Right':np.zeros((6,2,28))}
                
        for pp, pt in enumerate(self.pts):
            print(pt)
            for side in ['Left','Right']:
            #RIGHT NOW THIS DOES E1 FIRST then E3; so goes from lower number to higher number
            #NEEDS TO BE PLOTTED BACKWARDS
                rec_Zs[side][pp,:,:] = np.vstack((self.xs[pt][side][on_t_es[pp][0]-1][0:28],self.xs[pt]['Left'][on_t_es[pp][0]+1][0:28]))
        
        
                rec_Zs[side][rec_Zs[side] > 4000] = np.nan
        
        self.rec_Zs = rec_Zs
            
    def load_pt_OnT_rec(self):
        pt_recEs = np.array((6,2,2))
        pt_recOnTs = np.array([(ptont[pp][ss]+1,ptont[pp][ss]-1) for pp in range(len(self.do_pts))])

    def dynamics_measures(self):
        #plot the *diff* between weeks
        plt.figure()
        for ss,side in enumerate(['Left','Right']):
            plt.subplot(3,2,(ss+1)+0)
            plt.plot((self.Zdiff[side]))
            
            plt.subplot(3,2,(ss+1)+2)
            abs_Zdiff = np.abs(np.diff(self.Zdiff[side],axis=0))
            plt.plot(abs_Zdiff,alpha=0.2)
            plt.plot(np.nanmean(abs_Zdiff,axis=1),color='black')
        
        #plot the histogram of impedances, all of them
        plt.figure()
        for ss,side in enumerate(['Left','Right']):
            plt.subplot(1,2,ss+1)
            #stack everything together
            side_stack = np.vstack((self.rec_Zs[side][:,0,:],self.rec_Zs[side][:,1,:])).reshape(-1,1)
            side_stack = side_stack[~np.isnan(side_stack)]
            plt.hist(side_stack)
            plt.vlines(np.median(side_stack),0,100)
            print(side + ' ' + str(np.median(side_stack)))
            
    def gen_Zdiff(self):
        self.Zdiff = {'Left':[],'Right':[]}
        for side in ['Left','Right']:
            self.Zdiff[side] = np.squeeze(self.rec_Zs[side][:,1,:].T - self.rec_Zs[side][:,0,:].T)
        

    def plot_recZs(self):
        plt.figure()
        for ss,side in enumerate(['Left','Right']):
            plt.subplot(3,2,(ss+1)+0)
            plt.plot(self.rec_Zs[side][:,0,:].T)
            plt.ylim((500,2000))
            
            plt.subplot(3,2,(ss+1)+2)
            plt.plot(self.rec_Zs[side][:,1,:].T)
            plt.ylim((500,2000))
            
            plt.subplot(3,2,(ss+1)+4)
            plt.plot(self.rec_Zs[side][:,1,:].T - self.rec_Zs[side][:,0,:].T)
            plt.legend(self.pts)
            plt.ylim((-600,600))
        

#%%

print('Doing Zs')
Z_lib = Z_class()         
Z_lib.get_recZs()
#%%
Z_lib.plot_recZs()
Z_lib.dynamics_measures()
          
        
        
        
        
        
    #%%
            
#Deprecated class
class DEPRImpedances:
    def __init__(self,do_pts = ['DBS901','DBS903','DBS905','DBS906','DBS907','DBS908']):
        #load in the main file
        fname = '/home/virati/Dropbox/projects/Research/MDD-DBS/Data/Impedances/MDT_Z_Table.csv'
        Z_table = bringin_Zs_CSV(fname)
        
        electrodes = nestdict()
        
        electrodes['Left'] = np.zeros((4,48,6))
        electrodes['Right'] = np.zeros((4,48,6))
        
        for ch in range(4):
            electrodes['Left'][ch,:,:] = Z_table[ch::4,0::2]
            electrodes['Right'][ch,:,:] = Z_table[ch::4,1::2]
        
        #reshuffle the electrodes dictionary so we can actually work with a matrix
        big_Z_table = np.array([val for key,val in electrodes.items()])
        #self.electrodes = electrodes
        self.Z_table = big_Z_table
        
        
        self.do_pts = do_pts
        
        self.pt_OnT = self.load_pt_OnT_stim()
        self.clean_artifacts(electrodes)
        
    def load_etrode_map(self):
        #need a map of electrodes
        #so a dict
        etrode_map = {'Stim':0,'Rec':0}
        #in stim, we'll have a pt x side x etrodes(1)
        #in rec we'll have a pt x side x etrodes(2))
        etrode_map['Stim'] = self.load_pt_OnT_stim()
        etrode_map['Rec'] = self.load_pt_OnT_rec()
        
    
    def load_pt_OnT_stim(self):
        pt_OnT = [(2,1),(2,2),(2,1),(2,2),(1,1),(2,1)]
        #self.pt_OnT = {pt:pt_OnT[pp] for pp,pt in enumerate(self.do_pts)}
        return pt_OnT
    
    def load_pt_OnT_rec(self):
        pt_recEs = np.array((6,2,2))
        pt_recOnTs = np.array([(ptont[pp][ss]+1,ptont[pp][ss]-1) for pp in range(len(self.do_pts))])
        
        
    def clean_artifacts(self,etrode_dict):
        self.Z_table[self.Z_table > 35000] = np.nan
            
    #plot ALL zs forall contacts
    def plot_Zs(self,profile='Active'):
        
        #we want a plot that has all the patients Zs for the entire timecourse, on left and right, for all contacts
        #4x2 subplots
        
        plt.figure()
        for cc,chann in enumerate(range(3,-1,-1)):
            for ss,side in enumerate(['Left','Right']):
                plt.subplot(4,2,2*(cc)+(ss+1))
                plt.plot(self.Z_table[ss,chann,:28,:].squeeze(),alpha=0.2)
                plt.title('Channel e' + str(chann + ss*8))
            
                plt.plot(np.nanmean(self.Z_table[ss,chann,:28,:],axis=1),color='black')
                
        plt.suptitle('Electrode impedances over time')
    
    def plot_chann_map(self,direction='stim'):
        for pt in self.do_pts:
            pass
        
    def plot_Zdiff(self):
        self.pt_recEs = nestdict()
        ptont = self.pt_OnT
        plt.figure()
        
        z_both = np.zeros((2,2,28,6))
        
        abs_diff = np.zeros((2,28,6))
    
        for ss,side in enumerate(['Left','Right']):
            #we will only focus on two channels->Above and Below for each patient
            pt_recOnTs = np.array([(ptont[pp][ss]+1,ptont[pp][ss]-1) for pp in range(len(self.do_pts))])
            #let's see if this mask works....
            
            for pp,pt in enumerate(self.do_pts):
                z_both[:,0,:,pp] = self.Z_table[:,pt_recOnTs[pp,0],:28,pp]
                z_both[:,1,:,pp] = self.Z_table[:,pt_recOnTs[pp,1],:28,pp]
                
            
            plt.subplot(3,2,1+ss)
            plt.plot(z_both[ss,0,:,:])
            plt.title('Top Recording Electrode')
            
            plt.subplot(3,2,3+ss)
            plt.plot(z_both[ss,1,:,:])
            plt.title('Bottom Recording Electrode')
            
            plt.subplot(3,2,5+ss)
            #for pp,pt in enumerate(self.do_pts):
            abs_diff[ss,:,:] = np.abs(z_both[ss,1,:,:] - z_both[ss,0,:,:])
            plt.plot(abs_diff[ss,:,:])
            plt.legend(self.do_pts)
            plt.title('Recording Mismatch')
            plt.ylim((0,600))
            
            self.pt_recEs[side] = pt_recOnTs
             
        plt.suptitle('Impedance mismatch over time')
                
        
        # for ss,side in enumerate(['Left','Right']):
        #     zdiff = np.diff(self.electrodes[side][:,:,:],axis=1)
        #     for pp,pt in enumerate(self.do_pts):
        #         #make a matrix of the needed indices for recording electrodes
        #         pt_recOnT[pp,ss,:] = [ptont[pp][ss]-1,ptont[pp][ss]+1]
                
                
        #         plt.subplot(2,2,ss+1)
        #         plt.plot(zdiff[pt_recOnT[pp,ss,1],:29,pp].T,alpha=0.2)
        #         plt.title('Top Electrode')
                
        #         plt.subplot(2,2,ss+1+2*1)
        #         plt.plot(zdiff[pt_recOnT[pp,ss,0],:29,pp].T,alpha=0.2)
        #         plt.title('Bottom Electrode')
                
        #     plt.subplot(2,2,ss+1)
            
            
        #     #Make our mean matrix
        #     all_pt_recs = 0
            
        #     plt.plot(np.nanmean(zdiff[ptont[pp][ss]+1,:29,:],axis=1).T)
            
        #     plt.subplot(2,2,ss+1+2*1)
        #     plt.plot(np.nanmean(zdiff[ptont[pp][ss]-1,:29,:],axis=1).T)
            
        #     plt.legend()
        # plt.suptitle('Derivative of Impedances')
         
def DEPRbringin_Zs_CSV(fname='/home/virati/Dropbox/projects/Research/MDD-DBS/Data/Impedances/MDT_Z_Table.csv'):
    df = pd.read_csv(fname,header=None)
    #Now need to reshape
    big_table = df.as_matrix()
    return big_table





#%%
#GM_map = Anatomy()
        
#Z_lib = Impedances()
#Z_lib.plot_Zs(profile='Active')
#Z_lib.plot_Zdiff()